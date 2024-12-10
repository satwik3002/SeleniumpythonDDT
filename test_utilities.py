import openpyxl

def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]  # Corrected method to access the sheet
    return sheet.max_row  # Return the number of rows

def getcolumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]  # Corrected method to access the sheet
    return sheet.max_column  # Return the number of columns

def readData(file, sheetName, row, column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]  # Corrected method to access the sheet
    return sheet.cell(row=row, column=column).value  # Return the cell value

def writeData(file, sheetName, row, column, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]  # Corrected method to access the sheet
    sheet.cell(row=row, column=column).value = data  # Write the value to the cell
    workbook.save(file)  # Save the workbook
